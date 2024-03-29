from datetime import datetime
from pathlib import Path
from typing import Iterator
from backend.flow_logic import get_passed_levels_stats, process_stats
from backend.types import MAX_LANGUAGE_LEVEL, AllAnalytics, LanguageLevel, StagesAnalytics, TopicSuccessData
from backend.models import ProgressStep, Question, User, UserAnalytics, db_session
import openpyxl
from openpyxl.utils import get_column_letter
from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive
from sqlalchemy import Integer, case, func


def generate_users_results_export_data() -> Iterator[tuple]:
    for user in db_session.query(User):
        stats = get_passed_levels_stats(user.id)
        finished_level, _ = process_stats(stats)
        if finished_level is None:
            continue  # Export only users that have finished

        finished_level_str, recommended_group_str = '-', '-'
        if finished_level > LanguageLevel.A0:
            finished_level_str = str(finished_level)
        if finished_level < MAX_LANGUAGE_LEVEL:
            recommended_group_str = str(finished_level + 1)
        yield (
            user.id,
            user.full_name,
            user.email,
            finished_level_str,
            recommended_group_str,
            f'http://127.0.0.1:5000/results/{user.uuid}'
        )


def export_users_results_to_file(filepath: Path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Результаты прохождения теста'

    worksheet.append((
        'Id',
        'Полное имя',
        'Электронная почта',
        'Текущий уровень',
        'Рекомендуемая группа',
        'Подробные результаты',
    ))

    # Write the data
    for user_data in generate_users_results_export_data():
        worksheet.append(user_data)

    # Format links
    for row_idx in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=6)
        cell.hyperlink = cell.value

    # Resize columns
    for column_cells in worksheet.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            worksheet.column_dimensions[new_column_letter].width = new_column_length*1.23

    workbook.save(filepath)


def upload_file_to_google_drive(filepath: Path):
    settings = {
        'client_config_backend': 'service',
        'service_config': {
            'client_json_file_path': 'google-credentials.json',
        }
    }
    gauth = GoogleAuth(settings=settings)
    gauth.ServiceAuth()
    drive = GoogleDrive(gauth)
    google_file = drive.CreateFile({
        'parents': [{'kind': 'drive#fileLink', 'id': '1mHxFEdiXN4_4rZ9MXzjq1drdP6IRMtmM'}],
        'title': filepath.name,
    })
    google_file.SetContentFile(str(filepath))
    google_file.Upload()


def export_users_results_and_upload_to_google_drive():
    export_dir = Path(__file__).resolve().parent.parent / 'export_data'
    export_dir.mkdir(exist_ok=True)
    current_datetime_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = f'Экспорт результатов {current_datetime_str}.xlsx'
    filepath = export_dir / filename
    export_users_results_to_file(filepath)
    upload_file_to_google_drive(filepath)


def calculate_all_analytics() -> AllAnalytics:
    page_opened_count = db_session.query(UserAnalytics).count()
    started_the_test_count = db_session.query(User).count()
    finishsed_users_count = 0
    for user in db_session.query(User):
        finished_level, _ = process_stats(get_passed_levels_stats(user.id))
        if finished_level is not None:
            finishsed_users_count += 1

    level_determination_func = case((User.choosed_dont_know_level, LanguageLevel.A0), else_=User.start_level)
    start_level_distribution_query = db_session.query(
        level_determination_func,
        func.count(User.id),
    ).group_by(
        User.start_level,
    ).group_by(
        User.choosed_dont_know_level,
    ).order_by(
        level_determination_func
    )

    per_topic_query = db_session.query(
        Question.category,
        Question.topic_title,
        func.count(ProgressStep.question_id),
        func.sum(func.cast(ProgressStep.is_correct, Integer)),
    ).join(ProgressStep).filter(
        ProgressStep.answer.isnot(None),
    ).group_by(
        Question.category,
        Question.topic_title,
    ).order_by(
        Question.category,
        Question.topic_title,
    )

    topics_success = []
    for category, topic_title, questions_count, correct_answers_count in per_topic_query:
        topics_success.append(TopicSuccessData(
            category=category,
            topic_title=topic_title,
            questions_count=int(questions_count),
            correct_answers_count=int(correct_answers_count),
        ))

    all_analytics = AllAnalytics(
        stages_analytics=StagesAnalytics(
            opened_the_page_percentage=int((page_opened_count - started_the_test_count) / page_opened_count * 100),
            started_the_test_percentage=int((started_the_test_count - finishsed_users_count) / page_opened_count * 100),
            finished_the_test_percentage=int(finishsed_users_count / page_opened_count * 100),
        ),
        start_level_selection_distribution=[(level, int(count / started_the_test_count * 100)) for level, count in start_level_distribution_query],
        topics_success=topics_success
    )
    return all_analytics

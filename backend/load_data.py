from pathlib import Path
import random
import shutil
import sys
from openpyxl import load_workbook

from __init__ import AnswerType, LanguageLevel, Question, QuestionCategory, create_app, db_session


ROOT_DIR_PATH = Path(__file__).resolve().parent.parent / 'test_data'
print(f'Scanning {ROOT_DIR_PATH}')

ANSWER_TYPE_MAPPING = {
    'Выбор одного варианта': AnswerType.SELECT_ONE,
    'Выбор нескольких вариантов': AnswerType.SELECT_MULTIPLE,
    'Заполнение пропуска': AnswerType.FILL_THE_BLANK,
}


def process_topic_file(level: LanguageLevel, category: QuestionCategory, topic_file: Path) -> list[Question]:
    if not topic_file.is_file():
        print(f'{topic_file} should be a file')
        sys.exit(1)

    questions = []
    is_media = category in (QuestionCategory.LISTENING, QuestionCategory.READING)
    print(f'Reading topic file {topic_file.name}')
    workbook = workbook = load_workbook(topic_file)
    for sheet in workbook.worksheets:
        if sheet.title not in ANSWER_TYPE_MAPPING:
            print(f'Unknown answer type {sheet.title}')
            sys.exit(1)
        answer_type = ANSWER_TYPE_MAPPING[sheet.title]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_values = [str(cell) for cell in row if cell is not None]
            filepath = None
            if is_media:
                filepath = row_values[0]
                row_values = row_values[1:]
            question_title = row_values[0]
            if answer_type == AnswerType.SELECT_ONE:
                correct_answer_value = row_values[1]
                other_answers = row_values[2:]
                all_answers = [correct_answer_value] + other_answers
                random.shuffle(all_answers)
                answer_options = ','.join(all_answers)
                correct_answer = all_answers.index(correct_answer_value)
            elif answer_type == AnswerType.SELECT_MULTIPLE:
                correct_answers_num = int(row_values[1])
                correct_answers_values = row_values[2:2 + correct_answers_num]
                other_answers = row_values[2 + correct_answers_num:]
                all_answers = correct_answers_values + other_answers
                random.shuffle(all_answers)
                answer_options = ','.join(all_answers)
                correct_answer = ','.join([str(all_answers.index(value)) for value in correct_answers_values])
            elif answer_type == AnswerType.FILL_THE_BLANK:
                answer_options = None
                correct_answer = row_values[1]

            questions.append(Question(
                level=level,
                category=category,
                topic_title=topic_file.stem,
                question_title=question_title,
                filepath=filepath,
                answer_type=answer_type,
                answer_options=answer_options,
                correct_answer=correct_answer,
            ))

    return questions


collected_questions: list[Question] = []
collected_file_paths: list[Path] = []
for level_dir in ROOT_DIR_PATH.iterdir():
    if not level_dir.is_dir():
        print(f'{level_dir} should be a directory')
        sys.exit(1)

    transformed_level_name = level_dir.name.replace('.', '_')
    if transformed_level_name not in LanguageLevel.__members__:
        print(f'Unknown level {level_dir.name}')
        sys.exit(1)
    level = LanguageLevel[transformed_level_name]
    print(f'Skanning level {level_dir.name}')
    for meta_category_dir in level_dir.iterdir():
        if not meta_category_dir.is_dir():
            print(f'{meta_category_dir} should be a directory')
            sys.exit(1)

        print(f'Skanning meta category {meta_category_dir.name} of level {level_dir.name}')
        if meta_category_dir.name not in ('Грамматика', 'Лексика', 'Восприятие'):
            print(f'Unknown meta category {meta_category_dir.name}')
            sys.exit(1)
        if meta_category_dir.name == 'Грамматика':
            category = QuestionCategory.GRAMMAR
            for topic_file in meta_category_dir.iterdir():
                collected_questions.extend(process_topic_file(level, category, topic_file))
        elif meta_category_dir.name == 'Лексика':
            category = QuestionCategory.VOCABULARY
            for topic_file in meta_category_dir.iterdir():
                collected_questions.extend(process_topic_file(level, category, topic_file))
        else:  # Восприятие
            dirs = list(meta_category_dir.iterdir())
            for dir_path in dirs:
                if dir_path.name == 'Аудирование':
                    category = QuestionCategory.LISTENING
                    collected_questions.extend(process_topic_file(level, category, dir_path / 'Вопросы.xlsx'))
                    collected_file_paths.extend((dir_path / 'Аудиофайлы').glob('*.mp3'))
                elif dir_path.name == 'Чтение':
                    category = QuestionCategory.READING
                    collected_questions.extend(process_topic_file(level, category, dir_path / 'Вопросы.xlsx'))
                    collected_file_paths.extend((dir_path / 'Тексты').glob('*.txt'))


print(f'Loaded {len(collected_questions)} questions')
print(f'Loaded {len(collected_file_paths)} files')


app = create_app()
with app.app_context():
    db_session.add_all(collected_questions)
    db_session.commit()
    print('Questions added to database')

media_directory = Path(__file__).resolve().parent.parent / 'backend' / 'media'
media_directory.mkdir(exist_ok=True)

for filepath in collected_file_paths:
    shutil.copy(filepath, media_directory / filepath.name)